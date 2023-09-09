# Загружаем необходимые библиотеки
import io
import os
import simplecrypt
import tempfile
import getpass
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_workbook

# Переход в основную папку программы
def_dir = os.chdir('..')
start_dir = os.chdir('bases')

# Блок начального меню
def start_menu():
	os.system('clear')
	list_start_dir = os.listdir()
	clear_list = [element.replace('.nfr', '') for element in list_start_dir]
	print("")
	print("\033[5;1;33;45m KeePassIc \033[0m")
	print("")
	print("\033[36mБазы паролей:",clear_list, "\033[0m")
	print("")
	print("\033[32m '1' --- Создать базу паролей \033[0m")
	print("")
	print("\033[31m '2' --- Удалить базу паролей \033[0m")	
	print("")
	print("\033[32m '3' --- Открыть базу паролей \033[0m")
	print("")
	print("\033[33m '4' --- Переименовать базу паролей \033[0m")
	print("")
	print("\033[33m '5' --- Изменить ключ шифрования базы паролей \033[0m")
	print("")
	print("\033[31m '6' --- Выйти из программы \033[0m")
	print("")
	
# Блок работы с начальным меню
def start_menu_do():
        try:
                doit = input("Выберите действие: ")
        except EOFError:
                os.system('clear')
                exit()
        if doit == '1' :
                add_pass_base_menu()  
        elif doit == '2' :
                del_pass_base_menu()        
        elif doit == '3' :
                open_pass_base_menu()
        elif doit == '4' :
                rename_base_pases()
        elif doit == '5' :
            cha_key_base_pase()
        elif doit == '6' :
                os.system('clear')
                exit()
        else:
                print("\033[31mТакой функции нет! Повторите попытку!\033[0m")                              
                start_menu_do()
		
# Блок меню создания базы паролей		
def add_pass_base_menu():
	try:
		add_pass_base_menu_do = input("\n\033[32mВведите имя базы паролей\033[0m\n\033[33m'1'---Вернуться назад\033[0m\n\033[31m'2'---Выйти\033[0m\nВвод: ")
	except EOFError:
		os.system('clear')
		exit()
	if add_pass_base_menu_do == '1':
		os.system('clear')
		running()                
	elif add_pass_base_menu_do == '2':
		os.system('clear')
		exit()
	elif add_pass_base_menu_do == '':
		print ("\033[31mИмя не может быть пустым\033[0m")
		add_pass_base_menu()
	elif os.path.isfile(add_pass_base_menu_do + ".nfr"):
		print("\033[31mБаза с таким именем уже есть! Пвоторите попытку\033[0m")
		add_pass_base_menu()
	else:	
		pas_input = getpass.getpass("Введите ключ шифрования: ")		
		if pas_input == '':
			print("\033[31mКлюч шифрования не может быть пустым\033[0m")
			add_pass_base_menu()
		wb = Workbook()
		ws = wb.active
		ws.title = 'empty'
		ws['A1']='Логин:'
		ws['B1']='empty'
		ws['A2']='Пароль:'
		ws['B2']='empty'
		ws['A3']='Доп.информация:'
		ws['B3']='empty'
		data = io.BytesIO()
		wb.save(data)
		data.seek(0)
		txt = data.read()
		pas = pas_input
		print("\n\033[32mПодождите идет создание базы паролей\033[0m")	
		cipher = simplecrypt.encrypt(pas, txt)
		data.close()		
		fin = open(add_pass_base_menu_do + '.nfr', 'wb')
		fin.write(cipher)
		fin.close()
		print("")
		print("\033[32mБаза [" + add_pass_base_menu_do + "] создана\033[0m")
		sv_key = input("\033[33m'Enter'---Продолжить\n\033[31m'1'---Выйти\033[0m\nВвод: ")
		if sv_key == '1':
                        os.system('clear')
                        exit()
		running()

# Блок удаления базы паролей
def del_pass_base_menu():
	try:
		del_pass_base_menu_do = input("\n\033[32mВведите имя базы паролей\033[0m\n\033[33m'1'---Вернуться назад\033[0m\n\033[31m'2'---Выйти\033[0m\nВвод: ")
	except EOFError:
		os.system('clear')
		exit()
	if del_pass_base_menu_do == '2':
		os.system('clear')
		exit()
	elif del_pass_base_menu_do == '1':
		os.system('clear')
		running()
	elif del_pass_base_menu_do == '':
		print ("\033[31mИмя не может быть пустым\033[0m")
		del_pass_base_menu()
	else:
		try:
			os.remove(del_pass_base_menu_do + ".nfr")
		except FileNotFoundError: 
			print("\033[31mТакой базы не существует! Повторите попытку\033[0m")
			del_pass_base_menu()
		print("")
		print("\033[32mБаза [" + del_pass_base_menu_do + "] удалена\033[0m")
		del_key_pas = input("\033[33m'Enter'---Продолжить\033[0m\n\033[31m'1'---Выйти\033[0m\nВвод: ")
		if del_key_pas == '1':
		    os.system('clear')
		    exit()
		else:
			running()

#Блок переименования базы паролей
def rename_base_pases():
	rename_base_pases_do = input("\n\033[32mВведите имя базы паролей\033[0m\n\033[33m'1'---Вернуться назад\033[0m\n\033[31m'2'---Выйти\033[0m\nВвод: ")
	if rename_base_pases_do == '1':
		running()		
	elif rename_base_pases_do == '2':
		os.system('clear')
		exit()
	else:
		new_name_base_pases = input("Введите новое имя базы паролей: ")
		if new_name_base_pases == '':
			print("\033[31mИмя базы не может быть пустым\033[0m")
			rename_base_pases()
		try:
			fil = open(rename_base_pases_do + ".nfr", 'rb')
		except FileNotFoundError:
			print("\033[31mТакой базы паролей не существует\033[0m")
			rename_base_pases()
		inf = fil.read()
		fil.close()	
		crypt = open(new_name_base_pases + '.nfr', 'wb')
		crypt.write(inf)
		crypt.close()
		os.remove(rename_base_pases_do + ".nfr")
		end_in = input("\033[32m\nБаза паролей переименована\033[0m\n\033[33m'Enter'---Продолжить \n'1'---Выйти\033[0m\nВвод: ")
		if end_in == '1' :
			os.system('clear')
			exit()
		else :		
			running()

#Блок изменения ключа шифрования базы паролей
def cha_key_base_pase():
	menu = input("\n\033[32mВведите имя базы паролей\033[0m\n\033[33m'1'---Вернуться назад\033[0m\n\033[31m'2'---Выйти\033[0m\nВвод: ")
	if menu == '1':
		running()		
	elif menu == '2':
		os.system('clear')
		exit()
	else:
		check_key_base_pase = getpass.getpass("Введите ключ шифрования базы паролей: ")
		if check_key_base_pase == '':
			print("\033[31mКлюч шифрования не может быть пустым\033[0m")
			os.chdir('..')
			os.chdir('bases')
			cha_key_base_pase()
		try:	
			fil = open(menu + ".nfr", 'rb')
		except FileNotFoundError:
			print("\033[31mТакой базы не существует\033[0m")
			os.chdir('..')
			os.chdir('bases')
			cha_key_base_pase()
		inf = fil.read()
		fil.close()
		pas = check_key_base_pase
		print("\n\033[32mПодождите идет проверка ключа базы паролей\033[0m\n")	
		try:
			cipher_open_base = simplecrypt.decrypt(pas,inf)
		except simplecrypt.DecryptionException:
			print("\033[31mНеправильный ключ базы паролей\033[0m")
			os.chdir('..')
			os.chdir('bases')
			cha_key_base_pase()
		new_key_base_pase = getpass.getpass("Введите новый ключ шифрования базы паролей:")
		if new_key_base_pase == '':
			print("\033[31mКлюч шифрования не может быть пустым\033[0m")
			os.chdir('..')
			os.chdir('bases')
			cha_key_base_pase()
		new_pas = new_key_base_pase
		new_text = cipher_open_base
		print ("\033[32m\nПодождите идет изменение ключа базы паролей\033[0m")
		cipher_new_base = simplecrypt.encrypt(new_pas,new_text)	
		crypt = open(menu + '.nfr', 'wb')
		crypt.write(cipher_new_base)
		crypt.close()
		print ("\033[32m\nКлюч изменен")
		sv_key = input("\033[33m'Enter'---Продолжить\n\033[31m'1'---Выйти\033[0m\nВвод: ")
		if sv_key == '1':
			os.system('clear')
			exit()
		else:
			os.chdir('..')
			os.chdir('bases')
			running()
	
# Блок открытия базы паролей	
def open_pass_base_menu():
	global open_pass_base_menu_do
	try:
		open_pass_base_menu_do = input("\n\033[32mВведите имя базы паролей\033[0m\n\033[33m'1'---Вернуться назад\033[0m\n\033[31m'2'---Выйти\033[0m\nВвод: ")
	except EOFError:
		os.system('clear')
		exit()
	if open_pass_base_menu_do == '2':
		os.system('clear')
		exit()
	elif open_pass_base_menu_do == '1':
		os.system('clear')
		running()
	elif open_pass_base_menu_do == '':
		print ("\033[31mИмя не может быть пустым\033[0m")
		open_pass_base_menu()
	else:
		global key_pass_base
		key_pass_base = getpass.getpass("Введите ключ шифрования: ")
		work_with_pass_base()
		
# Блок работы с базой паролей
def work_with_pass_base():
	global base_pase_sheets
	global base_pase
	try:
		fil = open(open_pass_base_menu_do + ".nfr", 'rb')
	except FileNotFoundError:
			print("\033[31mТакой базы не существует! Повторите попытку\033[0m")
			os.chdir('..')
			os.chdir('bases')
			open_pass_base_menu()	
	inf = fil.read()
	fil.close()
	pas = key_pass_base
	print("\n\033[32mПодождите идет открытие базы паролей\033[0m\n")
	try:	
		cipher_open_base = simplecrypt.decrypt(pas,inf)
	except simplecrypt.DecryptionException:
		print("\033[31mКлюч шифрования не верен!\033[0m")
		os.chdir('..')
		os.chdir('bases')
		open_pass_base_menu()
	data = io.BytesIO()
	data.write(cipher_open_base)
	data.seek(0)
	base_pase = openpyxl.load_workbook(data)
	base_pase_sheets = base_pase.sheetnames
	os.system('clear')
	data.close()
	print("\033[36mАккаунты: ", base_pase_sheets, "\033[0m") 
	try:
		work_with_pass_base_do = input("\033[32m'1'---Посмотреть аккаунт\n\033[33m'2'---Добавить аккаунт\033[0m\n\033[31m'3'---Удалить аккаунт\n\033[33m'4'---Вернуться назад\033[0m\n\033[31m'5'---Выйти\033[0m\nВвод: ")
	except EOFError:
    		os.system('clear')
    		exit()

	if work_with_pass_base_do =='1':
		work_with_acc()	

	elif work_with_pass_base_do =='2':
		create_acc()

	elif work_with_pass_base_do =='3':
		del_acc()

	elif work_with_pass_base_do =='4':
		os.system('clear')
		data.close()
		running()

	elif work_with_pass_base_do == '5':
		os.system('clear')
		data.close()
		exit()
	
	else:
		print("\033[31mТакой функции нет! Повторите попытку\033[0m")
		work_with_pass_base()	

#Блок просмотра аккаунта
def work_with_acc():
	global name_acc_do
	name_acc_do = input("\n\033[32mВведите имя аккаунта\n\033[33m'1'---Вернуться назад \033[0m\nВвод: ")
	if name_acc_do == '1':
	    work_with_pass_base()
	try:
		ws = base_pase[name_acc_do]
	except KeyError:
		print("\033[31mТакого аккаунта не существует\033[0m")
		work_with_acc()
	ws1 = ws['A1'].value
	ws2 = ws['B1'].value
	ws3 = ws['A2'].value
	ws4 = ws['B2'].value
	ws5 = ws['A3'].value
	ws6 = ws['B3'].value
	print("")
	print("\033[32mИмя аккаунта:", name_acc_do)
	print("")	
	print(ws1, ws2)
	print(ws3, ws4)
	print(ws5, ws6, "\033[0m")
	print("")
	variants = input("\033[31m'1'---Изменить аккаунт\n\033[33m'Enter'---Вернуться назад\033[0m\nВвод: ")
	if variants == '1':
		chan_acc()
	else:
		work_with_pass_base()
	
#Блок удаления аккаунта
def del_acc():
	if len(base_pase_sheets) == 1:
		print("\n\033[31mХотя бы один аккаунт должен быть в базе\033[0m")
		work_with_pass_base()
	try:
		del_acc_input = input("\n\033[32mВведите имя удаляемого аккаунта\n\033[33m'1'---Вернуться назад\033[0m\nВвод: ")
	except EOFError:
    		os.system('clear')
    		exit()
	try:
		if del_acc_input == '1':
			work_with_pass_base()
		else:
			targ = base_pase[del_acc_input]	
			base_pase.remove(targ)		
			data = io.BytesIO()
			sv2 = base_pase.save(data)
			data.seek(0)
			txt = data.read()
			pas = key_pass_base
			print("\n\033[32mПодождите идет удаление аккаунта\033[0m")	
			cipher = simplecrypt.encrypt(pas, txt)
			data.close()
			fin = open(open_pass_base_menu_do + '.nfr', 'wb')
			fin.write(cipher)
			fin.close()
	except KeyError:
		print("\033[31mТакого аккаунта не существует\033[0m")
		del_acc()
	work_with_pass_base()

#Блок создания аккаунта
def create_acc():
	menu_select = input("\n\033[33m'1'---Вернуться назад\n'Enter'---Продолжить\033[0m\nВвод: ")
	if menu_select == '1':
		work_with_pass_base()
	cr_acc_name = input("\nВведи имя аккаунта: ")
	if cr_acc_name == '':
		print("\033[31mИмя аккаунта не может быть пустым\033[0m")
		create_acc()
	cr_acc_log = input("\nВведи логин: ")
	cr_acc_pas = input("Введи пароль: ")
	cr_acc_add = input("Введи доп. информацию: ")
	base_pase.create_sheet(cr_acc_name)			
	ws = base_pase[cr_acc_name]
	ws['A1']='Логин:'
	ws['B1']=cr_acc_log
	ws['A2']='Пароль:'
	ws['B2']=cr_acc_pas
	ws['A3']='Доп.информация:'
	ws['B3']=cr_acc_add
	data = io.BytesIO()
	sv2 = base_pase.save(data)
	data.seek(0)
	txt = data.read()
	pas = key_pass_base
	print("\n\033[32mПодождите идет сохранение аккаунта\033[0m")	
	cipher = simplecrypt.encrypt(pas, txt)
	data.close()
	fin = open(open_pass_base_menu_do + '.nfr', 'wb')
	fin.write(cipher)
	fin.close()
	work_with_pass_base()
	
#Блок изменения аккаунта
def chan_acc():
	select_vr = input("\n\033[31m'1'---Изменить имя аккаунта\n\n'2'---Изменить логин\n'3'---Изменить пароль\n'4'---Изменить Доп.Информацию\n\n\033[33m'Enter'---Вернуться назад\033[0m\nВвод: ")
	if select_vr =='1':
		chan_acc_name()
	elif select_vr =='2':
		chan_acc_log()
	elif select_vr =='3':
		chan_acc_pas()
	elif select_vr =='4':
		chan_acc_add()
	else:
		work_with_pass_base()

#Блок изменения имени аккаунта
def chan_acc_name():
	new_name_acc = input("\nВведите новое имя аккаунта: ")
	if new_name_acc == '':
		print("\033[31mИмя аккаунта не может быть пустым\033[0m")
		chan_acc()
	ws = base_pase[name_acc_do]
	ws.title = new_name_acc
	data = io.BytesIO()
	sv2 = base_pase.save(data)
	data.seek(0)
	txt = data.read()
	data.close()
	pas = key_pass_base
	print("\n\033[32mПодождите идет сохранение изменений\033[0m")	
	cipher = simplecrypt.encrypt(pas, txt)
	fin = open(open_pass_base_menu_do + '.nfr', 'wb')
	fin.write(cipher)
	fin.close()
	work_with_pass_base()
	
#Блок изменения логина аккаунта
def chan_acc_log():
	new_name_login = input("\nВведите новый логин: ")
	ws = base_pase[name_acc_do]
	ws['B1']=new_name_login
	data = io.BytesIO()
	sv2 = base_pase.save(data)
	data.seek(0)
	txt = data.read()
	data.close()
	pas = key_pass_base
	print("\n\033[32mПодождите идет сохранение изменений\033[0m")	
	cipher = simplecrypt.encrypt(pas, txt)
	fin = open(open_pass_base_menu_do + '.nfr', 'wb')
	fin.write(cipher)
	fin.close()
	work_with_pass_base()

#Блок изменения пароля аккаунта
def chan_acc_pas():
	new_name_pas = input("\nВведите новый пароль: ")
	ws = base_pase[name_acc_do]
	ws['B2']=new_name_pas
	data = io.BytesIO()
	sv2 = base_pase.save(data)
	data.seek(0)	
	txt = data.read()
	data.close()
	pas = key_pass_base
	print("\n\033[32mПодождите идет сохранение изменений\033[0m")	
	cipher = simplecrypt.encrypt(pas, txt)
	fin = open(open_pass_base_menu_do + '.nfr', 'wb')
	fin.write(cipher)
	fin.close()
	work_with_pass_base()

#Блок изменения Доп.информации аккаунта		
def chan_acc_add():
	new_name_add = input("\nВведите новую Доп.информацию: ")
	ws = base_pase[name_acc_do]
	ws['B3']=new_name_add
	data = io.BytesIO()
	sv2 = base_pase.save(data)
	data.seek(0)
	txt = data.read()
	data.close()
	pas = key_pass_base
	print("\n\033[32mПодождите идет сохранение изменений\033[0m")	
	cipher = simplecrypt.encrypt(pas, txt)
	fin = open(open_pass_base_menu_do + '.nfr', 'wb')
	fin.write(cipher)
	fin.close()
	work_with_pass_base()

# Блок запуска программы(В самый конец кода)
def running():
	start_menu()
	start_menu_do()
	
# Запуск программы	
running()
